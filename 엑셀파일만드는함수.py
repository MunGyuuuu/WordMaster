import openpyxl
from openpyxl import Workbook, load_workbook
from tkinter import filedialog
import random


def copy_excel_data(original_file_path):
    # 엑셀 파일 경로 설정
     
    new_file_path = 'selected_vocabulary.xlsx'

    # 엑셀 파일 불러오기
    workbook = openpyxl.load_workbook(original_file_path)
    sheet = workbook.active

    # 단어와 뜻을 저장할 리스트
    word_meaning_list = []

    # 엑셀 파일에서 단어와 뜻 읽어오기
    for row in sheet.iter_rows(min_row=2, values_only=True):  # 첫 번째 행은 헤더이므로 제외
        word, meaning = row[0], row[1]  # 첫 번째 열에 있는 단어와 두 번째 열에 있는 뜻 추출
        word_meaning_list.append((word, meaning))  # 단어와 뜻을 튜플 형태로 리스트에 추가

    # 단어와 뜻이 함께 저장될 리스트
    selected_word_meaning = random.sample(word_meaning_list, k=60)

    # 새로운 엑셀 파일 생성
    new_workbook = openpyxl.Workbook()
    new_sheet = new_workbook.active

    # 추출된 단어와 뜻을 새로운 엑셀 파일에 쓰기
    for idx, (word, meaning) in enumerate(selected_word_meaning, start=1):
        if idx <= 30:
            new_sheet.cell(row=2*idx-1, column=1, value=word)
            new_sheet.cell(row=2*idx-1, column=2, value=meaning)
        else:
            new_sheet.cell(row=2*idx-61, column=3, value=word)
            new_sheet.cell(row=2*idx-61, column=4, value=meaning)
    # 새로운 엑셀 파일 저장
    new_workbook.save(filename=new_file_path)

    # 엑셀 파일 닫기
    workbook.close()
    new_workbook.close()

    print(f"추출된 60개의 단어와 뜻이 {new_file_path}에 저장되었습니다.")
