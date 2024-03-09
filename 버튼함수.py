import tkinter as tk
from tkinter import filedialog
from openpyxl import Workbook, load_workbook
from os.path import basename

button_data = []  # 버튼 데이터를 저장할 리스트

check_password = '0000'

# 엑셀 파일에서 버튼 데이터 불러오기
def load_button_data():
    try:
        wb = load_workbook("savefilenameandfilepath.xlsx")  # 워크북 로드
        ws = wb.active  # 활성 시트 선택

        for row in ws.iter_rows(values_only=True):
            button_data.append(row)

    except FileNotFoundError:
        # 파일이 없는 경우, 새로운 엑셀 파일 생성
        save_button_data()

# 처음 창 생성
def create_initial_window():
    # 버튼 데이터 로드
    load_button_data()

    root_word = tk.Tk()
    root_word.title("단어장 선택하는 창")
    root_word.geometry("200x300")

    label_word = tk.Label(root_word, text="원하시는 단어장을 선택해주세요")
    label_word.pack(fill="x", padx=10, pady=3)

    add_button = tk.Button(root_word, text="버튼 추가하기", command=add_button_root)
    add_button.pack()

    for name, filepath in button_data:
        button = tk.Button(root_word, text=name, command=lambda path=filepath: open_excel_file(path))
        button.pack(fill="x", padx=10, pady=3)

    root_word.mainloop()


# 비밀번호 확인하는 함수
def verify_password(password):
    return password == check_password

# 버튼 추가 창 열기
def add_button_root():
    # 비밀번호 확인 창 띄우기
    password_check_window = tk.Tk()
    password_check_window.title("비밀번호 확인")
    password_check_window.geometry("200x100")

    label_password = tk.Label(password_check_window, text="비밀번호를 입력하세요:")
    label_password.pack()

    entry_password = tk.Entry(password_check_window, show="*")  # 입력 내용 감춤
    entry_password.pack()

    # 확인 버튼
    confirm_button = tk.Button(password_check_window, text="확인", command=lambda: verify_and_continue(entry_password.get(), password_check_window))
    confirm_button.pack()

    password_check_window.mainloop()

# 비밀번호 확인 후 버튼 추가 창 열기
def verify_and_continue(password, window):
    if verify_password(password):
        window.destroy()  # 비밀번호 확인 창 닫기
        open_button_add_window()  # 버튼 추가 창 열기
    else:
        # 잘못된 비밀번호를 입력한 경우 메시지 출력
        tk.messagebox.showerror("오류", "잘못된 비밀번호입니다.")
        
# 버튼 추가 창 열기
def open_button_add_window():
    root_addbutton = tk.Tk()
    root_addbutton.geometry("300x150")
    root_addbutton.title("버튼 추가하는 창")

    # Entry 위젯 생성
    label_addbutton_name = tk.Label(root_addbutton, text="파일 이름")
    label_addbutton_name.pack()

    textbox_addbutton_name = tk.Entry(root_addbutton)
    textbox_addbutton_name.pack()

    label_addbutton_filepath = tk.Label(root_addbutton, text="파일 경로")
    label_addbutton_filepath.pack()

    textbox_addbutton_filepath = tk.Entry(root_addbutton)
    textbox_addbutton_filepath.pack()

    # 파일 선택 버튼
    select_file_button = tk.Button(root_addbutton, text="파일 선택", command=lambda: get_file_path(textbox_addbutton_filepath, textbox_addbutton_name))
    select_file_button.pack()

    summit_button = tk.Button(root_addbutton, text="제출", command=lambda: summit_check_button(textbox_addbutton_name.get(), textbox_addbutton_filepath.get(), root_addbutton),create_initial_window)
    summit_button.pack()

# 창을 닫는 함수
def close_function(root_name):
    root_name.destroy()
    
# 접근 확인하는 함수
def access_check(password):
    if password == check_password:
        return 1
    else:
        return 0

# 제출 확인 함수
def summit_check_button(name, filepath, root_addbutton):
    # 버튼 데이터를 리스트에 추가
    button_data.append((name, filepath))
    print("버튼 추가 완료:", name)
    root_addbutton.destroy()  # 버튼 추가 창 닫기
    word_section()  # 단어장 선택 창 업데이트

# 단어장 선택하는 창 띄우기
def word_section():
    root_word = tk.Tk()
    root_word.title("단어장 선택하는 창")
    root_word.geometry("200x300")
    
    label_word = tk.Label(root_word, text="원하시는 단어장을 선택해주세요")
    label_word.pack(fill="x", padx=10, pady=3)
    
    add_button = tk.Button(root_word, text="버튼 추가하기", command=add_button_root)
    add_button.pack()

    for name, filepath in button_data:
        button = tk.Button(root_word, text=name, command=lambda path=filepath: open_excel_file(path))
        button.pack(fill="x", padx=10, pady=3)

    root_word.mainloop()

# 파일 대화 상자 열기
def open_file_dialog():
    filepath = filedialog.askopenfilename(title="Select a File", filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")])
    return filepath

# 파일 경로 가져오기
def get_file_path(entry_widget, name_widget):
    filepath = open_file_dialog()
    entry_widget.delete(0, tk.END)  # 이전 내용 삭제
    entry_widget.insert(tk.END, filepath)
    
    # 파일 이름 추출하여 자동완성
    filename = basename(filepath).split('.')[0]
    name_widget.delete(0, tk.END)
    name_widget.insert(tk.END, filename)

# 엑셀 파일에 버튼 데이터 저장
def save_button_data():
    wb = Workbook()  # 새 워크북 생성
    ws = wb.active  # 활성 시트 선택

    # 버튼 데이터를 시트에 저장
    for idx, (name, filepath) in enumerate(button_data, start=1):
        ws.cell(row=idx, column=1, value=name)
        ws.cell(row=idx, column=2, value=filepath)

    wb.save("savefilenameandfilepath.xlsx")  # 엑셀 파일로 저장

# 엑셀 파일에서 버튼 데이터 불러오기
def load_button_data():
    try:
        wb = load_workbook("savefilenameandfilepath.xlsx")  # 워크북 로드
        ws = wb.active  # 활성 시트 선택

        for row in ws.iter_rows(values_only=True):
            button_data.append(row)

    except FileNotFoundError:
        # 파일이 없는 경우, 새로운 엑셀 파일 생성
        save_button_data()
