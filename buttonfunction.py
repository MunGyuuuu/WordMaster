import os
import csv
from tkinter import filedialog, simpledialog, messagebox
from openpyxl import load_workbook
import tkinter as tk

def load_button_data():
    try:
        button_data = []
        wb = load_workbook("savebuttondata.xlsx")
        sheet = wb.active
        for row in sheet.iter_rows(values_only=True):
            button_data.append(row)
        return button_data
    except FileNotFoundError:
        return []

def show_success(value=None):
    
    if value:
        text = value+"되었습니다."
        messagebox.showinfo("성공",text)
    else:
        messagebox.showinfo("실패","실패하였습니다.")

def write_csv(data):
    script_dir = os.path.dirname(os.path.realpath(__file__))
    csv_data = os.path.join(script_dir, 'database.csv')
    path = script_dir +'\\'+'database.csv'
    with open(path, 'a', newline='') as file:
        writer = csv.writer(file)
        writer.writerow(data)
        



def summit_add_button(file_path,button_name=None):
    data = []
    filename, extension = os.path.splitext(file_path)
    filename_only = os.path.basename(filename)
    button_name = filename_only
    write_csv(data)



