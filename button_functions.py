from tkinter import filedialog
from tkinter import messagebox
import tkinter as tk
from openpyxl import Workbook, load_workbook
import os
from os.path import basename
from password_dialog import show_password_dialog


current_dir = os.getcwd()
button_data = []
check_password = '0000'

# 엑셀 파일에서 버튼 데이터 불러오기
def load_button_data():
    try:
        wb = load_workbook("savefilenameandfilepath.xlsx")  # 워크북 로드
        ws = wb.active  # 활성 시트 선택

        for row in ws.iter_rows(values_only=True):
            button_data.append(row)

    except FileNotFoundError:
        # 파일이 없는 경우, 새로운 엑셀 파일 생성
        save_button_data()

# 최신화된 버튼 데이터를 엑셀 파일에 저장하기
def save_button_data():
    wb = Workbook()  # 새 워크북 생성
    ws = wb.active  # 활성 시트 선택

    # 버튼 데이터를 시트에 저장
    for idx, (name, filepath) in enumerate(button_data, start=1):
        ws.cell(row=idx, column=1, value=name)
        ws.cell(row=idx, column=2, value=filepath)

    # 현재 코드가 있는 파일과 같은 곳에 저장
    file_name = "savefilenameandfilepath.xlsx"
    file_path = os.path.join(current_dir, file_name)

    wb.save(file_path)  # 엑셀 파일로 저장

def add_word_list():
    # 비밀번호 확인
    password = show_password_dialog()
    if password != check_password:
        messagebox.showerror("Error", "Incorrect password")
        return

    # 파일 대화 상자 열기
    file_path = open_file_dialog()
    if not file_path:
        messagebox.showerror("Error", "No file selected")
        return

    # 파일 이름 추출
    file_name = os.path.basename(file_path)
    word_list_name = os.path.splitext(file_name)[0]

    # 버튼 데이터 추가
    button_data.append((word_list_name, file_path))

    # 엑셀 파일에 저장
    save_button_data()

    messagebox.showinfo("Success", "Word list added successfully")


# 파일 대화 상자 열기
def open_file_dialog():
    filepath = filedialog.askopenfilename(title="Select a File", filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")])
    return filepath

# 파일 경로 가져오기
def get_file_path(entry_widget, name_widget):
    filepath = open_file_dialog()
    entry_widget.delete(0, tk.END)  # 이전 내용 삭제
    entry_widget.insert(tk.END, filepath)

    # 파일 이름 추출하여 자동완성
    filename = basename(filepath).split('.')[0]
    name_widget.delete(0, tk.END)
    name_widget.insert(tk.END, filename)
